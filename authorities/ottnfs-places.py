# authorities/ottnfs-places.py

"""
Stage the Ottoman NFS Gazetteer (Kabadayı, Boykov, Sefer & Gerrits, 2022)
to the staged extract directory used by the rebuild / incremental pipeline.

Source: https://zenodo.org/records/7351936  (CC-BY 4.0)
  "Kabadayi_Boykov_Sefer_Gerrits_Ottoman_NFS_Gazetteer_..._16296_populated_places"
  16,296 populated places geocoded from mid-19th-century Ottoman population
  registers (NFS.d. = *Nüfus Defterleri*), 1830-1849. Anatolia, SE Europe,
  Bulgaria. POINT geometries only (longitude/latitude columns), single .xlsx.

Output: ``{STAGED_BASE_DIR}/ofs/extract/places.jsonl``

ES indexing happens later via the incremental ``index_namespace`` path (this
authority is small + point-only, so it suits the single-namespace add workflow,
NOT a full rebuild cutover). This script never talks to Elasticsearch.

Verified against the real v1 .xlsx (Sheet1, 16,296 rows × 15 cols, headers in
HEADER_ALIASES below). register_date_in_Hicri is an integer Hijri year (e.g.
1256 AH ≈ 1840 CE); hijri_year_to_gregorian() converts it, falling back to the
1830-1849 register window midpoint. Re-run `--dump-headers` if a future Zenodo
version bumps the schema.
----------------------------------------------------------------------------

=== INCREMENTAL SINGLE-NAMESPACE ADD RUNBOOK (ns=ofs) ======================
POINT-ONLY SHORTCUT: ofs is point-only, so the geom_store/H3/ccode staging
chain (steps 2-3 below) is ALL NO-OPS and is SKIPPED — verified on the
2026-06-05 prod load. Points have has_geom=False / geom_ref=None (nothing goes
to the geom store; the global geom staging dir stays empty), helpers compute
h3_cover/h3_centroid INLINE during EXTRACT (from repr_point, which IS the real
H3 for a point), and ccodes are left unset (assigned spatially later). So index
straight from the `extract` stage. Keep steps 2-3 only for POLYGON authorities
(e.g. ukhc), where H3 must come from the real geom-store polygon.
Run on CRC unless noted; long Python is fine on `pitt`, never on a CRC login
node — use a compute node / Slurm.
Activate first:  source <conda>/etc/profile.d/conda.sh && conda activate whg && cd /vast/ishi/elastic

 0. FETCH + verify — auto-download the .xlsx (AUTHORITIES['ofs'] points at the
    stable Zenodo direct-asset endpoint), then confirm headers map cleanly:
      python -m processing.fetch_authorities -n ofs --age 0
      python -m authorities.ottnfs-places --dump-headers   # expect no UNMAPPED

 1. EXTRACT — stage places.jsonl (also computes h3_cover/h3_centroid inline):
      python -m authorities.ottnfs-places            # uses AUTHORITIES['ofs'] file

 2-3. SKIPPED for ofs (point-only — see POINT-ONLY SHORTCUT above). For a
    polygon authority you would instead run, IN ORDER:
      python -m processing.geom_store --merge --keep-staging      # before H3
      python -m processing.h3_stage   --run-id <RID> --namespace <ns>
      python -m processing.h3_merge   --run-id <RID> --namespace <ns>
      : > ${STAGED_BASE_DIR}/<ns>/ccode/places.ccode.jsonl        # empty pass-through
      python -m processing.ccode_merge --run-id <RID> --namespace <ns>
    (h3_stage/h3_merge/ccode_merge are manifest-driven — they need a run-id.)

 4. INDEX into the live `places` alias + augment toponyms. Source stage =
    `extract` for ofs (it already carries h3_cover; index_namespace's _missing_h3
    guard only trips on has_geom=True geoms, so points are exempt). MUST run ON
    pitt — prod ES is firewalled to pitt's localhost:9201; /vast/ishi/staged is
    shared so pitt sees the extract. pitt has no /ihome: use
    /home/gazetteer/miniconda/envs/whg/bin/python. NOTE: index_namespace builds
    Elasticsearch(es_host) with NO auth arg AND prints es_host, so do NOT embed
    the prod password in the URL (401 without, leaks with). Either pass a
    no-auth host on an unsecured port, or drive its functions with a basic_auth
    client (pw from /ix1/ishi/es/config/elastic.password) — see git log
    2026-06-05 for the driver snippet. Dry-run first (omit --execute):
      python -m processing.index_namespace --namespace ofs --source-stage extract --es-host <PROD>
      python -m processing.index_namespace --namespace ofs --source-stage extract --es-host <PROD> --execute
    (Appends place_id to toponym attestations + 'ofs' to namespaces, never
     overwrites embeddings. New toponyms are created WITHOUT embeddings → run the
     Symphonym backfill in step 5a so they become fuzzy/phonetic-searchable.)

 5a. EMBEDDING BACKFILL — new toponyms are created WITHOUT a Symphonym vector
    (exact/prefix-searchable, but not fuzzy/phonetic KNN). The toponyms index
    stores ONLY the `embedding` (128-d byte) — no IPA/PanPhon — and inference is
    name-only (Student model). Three phases, /vast-bridged (prod ES on pitt, GPU
    on CRC). Preferred input is the index-time emission (no ES re-scan):
      # at step 4, add: --emit-new-toponyms /vast/ishi/staged/ofs/backfill/new.jsonl
      # compute on a CRC GPU node (sbatch -M gpu --partition a100 --gres=gpu:1):
      python -m phonetics.inference.backfill_embeddings compute \
          --in /vast/ishi/staged/ofs/backfill/new.jsonl \
          --out /vast/ishi/staged/ofs/backfill/embeddings.jsonl --device cuda
      # index on pitt (embedding_version MUST match the index — currently 7):
      python -m phonetics.inference.backfill_embeddings index --es-host <PROD> \
          --in /vast/ishi/staged/ofs/backfill/embeddings.jsonl --embedding-version 7
    Or, post-hoc (also heals pre-existing gaps), run `export --namespace ofs`
    instead of --emit-new-toponyms to derive the set from ES (must_not exists
    embedding). ofs's 2026-06-05 load used this post-hoc path (28,643 toponyms,
    ~2s GPU compute, 8s index).

 5. AGGREGATES (feed the registry push):
      python -m processing.gazetteer_h3_coverage   --namespace ofs
      python -m processing.gazetteer_temporal_extent --namespace ofs

 6. TILES — bucket `ofs` already registered in generate_tiles._PER_NAMESPACE_BUCKETS.
    NB: `submit_tiles_slurm --only-bucket ofs` will NOT queue ofs here — its
    per-namespace gate requires the `aat_enrich` stage marked complete in a run
    manifest, which this incremental chain (ends at ccode_merge/final) never
    writes. So use the ONE-OFF direct invocation, which bypasses the manifest
    gate, generates just `ofs.mbtiles`, and auto-deploys (proxied via pitt).
    Must run on a CRC COMPUTE NODE — generate_tiles won't import on pitt
    (antimeridian dep). One-off sbatch (htc, small tier is ample for ~16k pts):

      cat > /tmp/tiles_ofs.sbatch <<'EOF'
      #!/bin/bash
      #SBATCH --job-name=whg-tiles-ofs
      #SBATCH --partition=htc
      #SBATCH --qos=htc-htc-s
      #SBATCH --nodes=1 --ntasks=1 --cpus-per-task=4 --mem=16G --time=01:00:00
      #SBATCH --output=/vast/ishi/elastic/logs/whg-tiles-ofs-%j.out
      #SBATCH --error=/vast/ishi/elastic/logs/whg-tiles-ofs-%j.err
      set -eo pipefail
      ulimit -n 65536
      source /ihome/ishi/stg135/miniconda3/etc/profile.d/conda.sh
      conda activate whg
      cd /vast/ishi/elastic
      python -u -m processing.generate_tiles --bucket ofs   # auto-deploys; add --no-deploy to dry-run
      EOF
      sbatch -M htc /tmp/tiles_ofs.sbatch

    Then register the tileset in the tileserver config.json (auto-deploy only
    ships the .mbtiles; it does NOT rewrite config). Safe single-bucket rewrite
    + restart + serving-verify + auto-rollback (run from pitt, NOT a login node):
      python -m processing.update_tileserver_config --bucket ofs --execute

 7. REGISTRY (LAST — preflight refuses unless the ofs tileset actually serves):
      python -m processing.push_gazetteer_inventory --namespace ofs

 8. Re-read the geom store for exact containment:
      es gateway-restart
    NB: NOT needed for a point-only ns like ofs — exact containment reads the
    geom store, which holds no point entries. Only run it for polygon adds.
===========================================================================
"""

import re
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook

from processing.helpers import (
    enrich_geometry,
    compute_h3_fields,
    select_h3_cover_geometry,
    write_staged_place_doc,
)
from processing.settings import DATA_DIR, AUTHORITIES

NAMESPACE = "ofs"  # Ottoman NFS gazetteer (rename freely; keep short per convention)
KAZA_NS = "ofs-kaza"  # pseudo-namespace for named (unresolved) admin parents

OFS_CONFIG = next((a for a in AUTHORITIES if a['namespace'] == NAMESPACE), None)

# The register window: 764 registers dated 1830-1849. Used as the place's
# attestation timespan and as the date fallback when a row's Hicri year is
# unparseable.
REGISTER_START, REGISTER_END = 1830, 1849
REGISTER_MID = 1840

# Map of normalised header -> our canonical key. Normalisation lowercases and
# strips non-alphanumerics, so "kaza in NFS.d." -> "kazainnfsd". The exact keys
# below are verified against the real v1 .xlsx (Sheet1, 15 cols); generic
# synonyms are kept as a defensive fallback should a future version rename.
HEADER_ALIASES = {
    # geometry                                          (cols 13-14)
    'latitude': 'lat', 'lat': 'lat', 'y': 'lat',
    'longitude': 'lon', 'lon': 'lon', 'long': 'lon', 'x': 'lon',
    # identity                                          (col 0)
    'populatedplaceid': 'place_id', 'placeid': 'place_id', 'id': 'place_id',
    # toponyms                                          (cols 6-8)
    'toponymtranscribedfromnfsd': 'translit', 'transcribedtoponym': 'translit',
    'toponymottomaninnfsd': 'ottoman', 'ottomantoponym': 'ottoman',
    'toponymmodern': 'modern', 'moderntoponym': 'modern',
    # admin hierarchy. Note there is NO plain "liva in NFS.d." column — liva is
    # only present as the 1848-yearbook standardisation. kaza appears twice:
    # the register's own spelling (col 3) and the 1848 standardisation (col 2).
    'kazainnfsd': 'kaza',                               # col 3 (register spelling)
    'kaza18481264': 'kaza_1848',                        # col 2 (1848 yearbook; 1264 AH)
    'liva18481264': 'liva_1848',                        # col 1 (1848 yearbook liva/sancak)
    'nahiyeinnfsd': 'nahiye',                           # col 4
    'divaninnfsd': 'divan',                             # col 5
    # provenance                                        (cols 9-12)
    'project': 'project',                               # UrbanOccupationsOETR / POPGEO_BG
    'nfsdregisternumber': 'reg_no', 'registernumber': 'reg_no',
    'doctype': 'doc_type',                              # mufassal / icmal
    'registerdateinhicri': 'reg_date', 'registerdate': 'reg_date',
}


def _norm(h):
    return re.sub(r'[^a-z0-9]', '', str(h).lower()) if h is not None else ''


def hijri_year_to_gregorian(cell):
    """Best-effort Hijri-year -> Gregorian-year. Returns int or None.

    Scrapes the first 3-4 digit run (the Hijri year, ~1246-1266 AH for this
    corpus) and applies the standard conversion G ≈ H*0.970224 + 621.5.
    """
    if cell is None:
        return None
    m = re.search(r'(1[0-9]{3}|[0-9]{3,4})', str(cell))
    if not m:
        return None
    h = int(m.group(1))
    # Heuristic: a value already in the Gregorian register window is passed
    # through (some rows may carry a converted date).
    if REGISTER_START - 5 <= h <= REGISTER_END + 5:
        return h
    g = int(round(h * 0.970224 + 621.5))
    return g if REGISTER_START - 30 <= g <= REGISTER_END + 30 else REGISTER_MID


def _slug(s):
    return re.sub(r'[^a-z0-9]+', '-', str(s).strip().lower()).strip('-')


def _toponym(name, lang, ts, toponyms, seen):
    name = (name or '').strip()
    if not name:
        return
    lst = f"{name}@{lang}"
    if lst in seen:
        return
    seen.add(lst)
    toponyms.append({'toponym_id': lst, 'timespans': ts})


def process_row(row):
    """Map one normalised {canonical_key: value} row to a place doc, or None."""
    place_raw = row.get('place_id')
    if not place_raw:
        return None
    place_id = f"{NAMESPACE}:{str(place_raw).strip()}"

    # --- geometry: points only -------------------------------------------
    geometry = None
    try:
        lon, lat = float(row['lon']), float(row['lat'])
        if -180 <= lon <= 180 and -90 <= lat <= 90 and not (lon == 0 and lat == 0):
            geometry = {'type': 'Point', 'coordinates': [lon, lat]}
    except (KeyError, TypeError, ValueError):
        pass

    # --- attestation timespan from the register date ---------------------
    year = hijri_year_to_gregorian(row.get('reg_date')) or REGISTER_MID
    timespans = [{'start': {'in': REGISTER_START}, 'end': {'in': REGISTER_END}}]

    # --- modern toponym: detect the placeholders -------------------------
    # ~797 rows carry the literal "vanished" in toponym_modern (a settlement
    # mapped from historical sources with no surviving modern counterpart);
    # one carries "n/a". These are NOT real names — strip them, and use
    # "vanished" to flag the place type. A trailing "?" on a real modern name
    # marks an uncertain identification; drop it for clean toponym matching.
    modern_raw = (row.get('modern') or '').strip()
    vanished = modern_raw.lower() == 'vanished'
    modern = '' if modern_raw.lower() in ('vanished', 'n/a', 'na') \
        else modern_raw.rstrip('? ').strip()

    # --- toponyms (multi-script — feeds Symphonym cross-script matching) --
    # ota = Ottoman Turkish (Arabic script). Modern/translit are Latin; the
    # region spans TR + BG + GR, so modern names are tagged @und (unknown)
    # rather than forcing @tr.
    toponyms, seen = [], set()
    _toponym(row.get('ottoman'), 'ota', timespans, toponyms, seen)
    _toponym(row.get('translit'), 'ota-Latn', timespans, toponyms, seen)
    modern_ts = [{'start': {'in': 2000}, 'end': {'in': 2025}}]
    _toponym(modern, 'und', modern_ts, toponyms, seen)

    if not toponyms:
        return None  # nothing searchable

    # Title: prefer the transcribed Latin form, then modern, then Ottoman.
    title = ((row.get('translit') or '').strip() or modern
             or (row.get('ottoman') or '').strip())

    geom_entry = enrich_geometry(geometry, timespans=timespans) if geometry else None
    place_doc = {
        'place_id': place_id,
        'title': title,
        'toponyms': toponyms,
        'geometries': [geom_entry] if geom_entry else [],
    }
    if geom_entry and geom_entry.get('repr_point'):
        rp = geom_entry['repr_point']
        h3_geom = select_h3_cover_geometry(geom_entry, geometry)
        h3c, h3cover = compute_h3_fields(rp['lon'], rp['lat'], h3_geom)
        if h3c:
            place_doc['h3_centroid'] = h3c
            place_doc['h3_cover'] = h3cover

    # --- type: single implicit AAT type ----------------------------------
    # Every record is a "populated place" -> AAT 300008347 "inhabited places".
    # The ~797 "vanished" settlements (no surviving modern counterpart) are
    # flagged via sourceLabel; `vanished` is set from the modern-column literal.
    place_doc['types'] = [{
        'identifier': 'aat:300008347',
        'label': 'ottnfs',
        'sourceLabel': 'vanished populated place (NFS.d.)' if vanished
                       else 'populated place (NFS.d.)',
    }]

    # --- admin parents as NAMED relations (NOT resolved geometries) -------
    # The hierarchy is liva → kaza → nahiye → divan. These are free-text names,
    # not IDs, so we synthesise stable pseudo-ids under `ofs-kaza:` to graph-link
    # siblings without falsely claiming a match to a real boundary in `places`.
    # Resolving them to OHM/GeoNames admin polygons is a separate, fuzzy step.
    # The kaza uses the 1848-yearbook standardisation when present (more
    # consistent across the corpus), falling back to the register spelling.
    kaza_name = (row.get('kaza_1848') or row.get('kaza') or '').strip()
    admin = (
        ('liva', (row.get('liva_1848') or '').strip()),
        ('kaza', kaza_name),
        ('nahiye', (row.get('nahiye') or '').strip()),
        ('divan', (row.get('divan') or '').strip()),
    )
    relations = [
        {
            'relation_type': 'within',
            'related_place_id': f"{KAZA_NS}:{lvl}:{_slug(name)}",
            'label': f"{lvl}: {name}",
            'timespans': timespans,
        }
        for lvl, name in admin if name
    ]
    if relations:
        place_doc['relations'] = relations

    # --- provenance / display extras (small per-namespace; mirrors the
    #     historical_county/parish precedent in indexvillaris) -------------
    #   kaza        = register spelling (col 3, "kaza in NFS.d.")
    #   kaza_1848   = 1848-yearbook standardisation (col 2)
    #   liva_1848   = 1848-yearbook liva/sancak (col 1)
    for src, dst in (('kaza', 'kaza'), ('kaza_1848', 'kaza_1848'),
                     ('liva_1848', 'liva_1848'), ('nahiye', 'nahiye'),
                     ('divan', 'divan'), ('reg_no', 'register_no'),
                     ('doc_type', 'register_type'), ('project', 'source_project')):
        val = (str(row[src]).strip() if row.get(src) not in (None, '') else '')
        if val:
            place_doc[dst] = val
    if year:
        place_doc['register_year'] = year

    # ccodes deliberately left unset — assigned spatially downstream.
    return place_doc


def _row_iter(xlsx_path, dump_headers=False):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    canon = [HEADER_ALIASES.get(_norm(h)) for h in header]
    if dump_headers:
        print("Headers found (raw -> normalised -> canonical):")
        for h in header:
            print(f"  {h!r:40} -> {_norm(h):24} -> {HEADER_ALIASES.get(_norm(h))}")
        unmapped = [h for h in header if HEADER_ALIASES.get(_norm(h)) is None]
        if unmapped:
            print(f"\n  UNMAPPED (extend HEADER_ALIASES): {unmapped}")
        return
    for raw in rows:
        yield {canon[i]: raw[i] for i in range(len(canon)) if canon[i]}


def stage_ottnfs_file(xlsx_path):
    print(f"Processing Ottoman NFS gazetteer: {xlsx_path}")
    if not Path(xlsx_path).exists():
        std = Path(DATA_DIR) / 'authorities' / NAMESPACE / Path(xlsx_path).name
        if std.exists():
            xlsx_path = std
        else:
            print(f"ERROR: File not found: {xlsx_path}")
            return

    staged, skipped, errors = 0, 0, 0
    start = datetime.now()
    for i, row in enumerate(_row_iter(xlsx_path)):
        if (i + 1) % 1000 == 0:
            print(f"\r  {i + 1} rows - staged: {staged}", end='', flush=True)
        try:
            doc = process_row(row)
            if not doc:
                skipped += 1
                continue
            write_staged_place_doc(namespace=NAMESPACE, doc=doc)
            staged += 1
        except Exception as e:
            print(f"\n  ERROR row {i}: {e}")
            errors += 1

    print(f"\n{'=' * 80}\nOTTOMAN NFS STAGING COMPLETE\n{'=' * 80}")
    print(f"Time: {(datetime.now() - start).seconds}s")
    print(f"Staged: {staged:,}\nSkipped: {skipped:,}\nErrors: {errors:,}")


if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser(description='Stage the Ottoman NFS gazetteer')
    p.add_argument('--file', help='Path to the .xlsx (defaults to configured file)')
    p.add_argument('--dump-headers', action='store_true',
                   help='Print column headers + normalisation and exit')
    args = p.parse_args()

    if args.file:
        xlsx = args.file
    elif OFS_CONFIG and OFS_CONFIG.get('files'):
        name = OFS_CONFIG['files'][0].get('name') or Path(OFS_CONFIG['files'][0]['url']).name
        xlsx = Path(DATA_DIR) / 'authorities' / NAMESPACE / name
    else:
        print("ERROR: no --file and no AUTHORITIES['ofs'] config")
        sys.exit(1)

    if args.dump_headers:
        list(_row_iter(xlsx, dump_headers=True))
        sys.exit(0)

    print(f"Ottoman NFS gazetteer (STAGING)\nFile: {xlsx}\n")
    stage_ottnfs_file(str(xlsx))
